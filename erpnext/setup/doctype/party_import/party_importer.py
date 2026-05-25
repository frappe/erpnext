# Copyright (c) 2026, Frappe Technologies Pvt. Ltd. and contributors
# For license information, please see license.txt

import csv
import io
from collections import defaultdict

import frappe
from frappe import _
from frappe.utils import cint, cstr, flt, get_datetime, getdate
from frappe.utils.csvutils import read_csv_content
from frappe.utils.xlsxutils import read_xls_file_from_attached_file, read_xlsx_file_from_attached_file

PARTY_IMPORT_FIELDS = {
	"Customer": [
		"party_code",
		"customer_name",
		"customer_type",
		"customer_group",
		"default_currency",
		"territory",
		"default_bank_account",
		"tax_id",
		"website",
	],
	"Supplier": [
		"party_code",
		"supplier_name",
		"supplier_type",
		"supplier_group",
		"default_currency",
		"tax_id",
		"website",
	],
	"Address": [
		"address_type",
		"address_line1",
		"address_line2",
		"city",
		"state",
		"pincode",
		"country",
		"phone",
		"email_id",
	],
	"Contact": ["first_name", "last_name", "gender", "designation", "department", "email_id", "phone"],
}

EXCLUDED_FIELDTYPES = {
	"Section Break",
	"Column Break",
	"HTML",
	"Button",
	"Fold",
	"Heading",
	"Tab Break",
	"Table",
	"Table MultiSelect",
	"Image",
	"Attach Image",
}

EXCLUDED_FIELDNAMES = {
	"naming_series",
	"amended_from",
	"amendment_date",
	"docstatus",
	"parent",
	"parentfield",
	"parenttype",
	"idx",
	"owner",
	"creation",
	"modified",
	"modified_by",
	"lft",
	"rgt",
	"old_parent",
	"default_price_list",
	"language",
	"disabled",
	"",
}

EXCLUDED_CHILD_TABLES = {"links", "terms", "companies", "sales_team", "portal_users"}


def _fieldname_for_label(doctype, label):
	if label == "Party Code":
		return "party_code"

	try:
		meta = frappe.get_meta(doctype)
		df = next((f for f in meta.fields if f.label == label), None)
		if not df:
			df = meta.get_field(label)
		return df.fieldname if df else None
	except Exception:
		return None


def _project_section(row, prefix):
	out = {}
	for key, val in row.items():
		if key.startswith(prefix):
			out[key[len(prefix) :]] = val
	return out


class PartyGroup:
	def __init__(self, party_code, party_doctype):
		self.party_code = party_code
		self.party_doctype = party_doctype
		self.party_row = None
		self.address_rows = []
		self.contact_rows = []
		self.row_indices = []

	def add_row(self, row):
		self.row_indices.append(row.get("_row_index", 0))

		party_data = _project_section(row, f"{self.party_doctype}::")
		if self.party_row is None:
			self.party_row = party_data

		addr_data = _project_section(row, "Address::")
		if addr_data.get("address_line1") or addr_data.get("city") or addr_data.get("pincode"):
			self.address_rows.append(addr_data)

		contact_data = _project_section(row, "Contact::")
		email_child = _project_section(row, "Contact Email::")
		phone_child = _project_section(row, "Contact Phone::")

		extra_emails = [e for e in (email_child.get("email_id"),) if e]
		extra_phones = [p for p in (phone_child.get("phone"),) if p]

		if any(contact_data.values()) or extra_emails or extra_phones:
			contact_row = dict(contact_data)
			if extra_emails:
				contact_row["_extra_emails"] = extra_emails
			if extra_phones:
				contact_row["_extra_phones"] = extra_phones
			self.contact_rows.append(contact_row)


def get_template_fields(party_type):
	party_doctype = "Supplier" if party_type == "Supplier" else "Customer"
	sections = []
	mandatory = set()
	party_mandatory = set()

	for dt in [party_doctype, "Address", "Contact"]:
		meta = frappe.get_meta(dt)
		fields = []

		for fieldname in PARTY_IMPORT_FIELDS.get(dt, []):
			if fieldname == "party_code":
				fields.append({"fieldname": "party_code", "label": "Party Code", "reqd": 1, "depends_on": ""})
				mandatory.add("party_code")
				party_mandatory.add("party_code")
				continue

			df = meta.get_field(fieldname)
			if not df:
				continue

			fields.append(
				{
					"fieldname": df.fieldname,
					"label": df.label or fieldname,
					"reqd": int(bool(df.reqd)),
					"depends_on": df.depends_on or "",
				}
			)
			if df.reqd:
				mandatory.add(fieldname)
				if dt == party_doctype:
					party_mandatory.add(fieldname)

		sections.append(
			{
				"doctype": dt,
				"label": dt,
				"fields": fields,
				"subsections": _get_child_table_sections(dt),
			}
		)

	return {
		"sections": sections,
		"mandatory_fieldnames": list(mandatory),
		"party_mandatory_fieldnames": list(party_mandatory),
	}


def _get_child_table_sections(doctype):
	meta = frappe.get_meta(doctype)
	subsections = []
	for df in meta.fields:
		if df.fieldtype != "Table" or df.hidden or df.fieldname in EXCLUDED_CHILD_TABLES:
			continue
		if not df.options:
			continue
		child_fields = _get_child_fields(df.options)
		if child_fields:
			subsections.append(
				{
					"doctype": df.options,
					"label": df.label or df.options,
					"parent_field": df.fieldname,
					"fields": child_fields,
				}
			)
	return subsections


def _get_child_fields(child_doctype):
	try:
		meta = frappe.get_meta(child_doctype)
	except Exception:
		return []
	fields = []
	for df in meta.fields:
		if df.fieldtype in EXCLUDED_FIELDTYPES or df.hidden:
			continue
		if df.fieldname in EXCLUDED_FIELDNAMES:
			continue
		if df.read_only and not df.reqd:
			continue
		fields.append(
			{
				"fieldname": df.fieldname,
				"label": df.label or df.fieldname,
				"reqd": int(bool(df.reqd)),
				"depends_on": df.depends_on or "",
			}
		)
	return fields


def _flatten_template_fields(party_type):
	meta_data = get_template_fields(party_type)
	fields = []
	for section in meta_data["sections"]:
		for f in section["fields"]:
			fields.append({"doctype": section["doctype"], "fieldname": f["fieldname"], "label": f["label"]})
		for sub in section.get("subsections", []):
			for f in sub["fields"]:
				fields.append(
					{
						"doctype": sub["doctype"],
						"fieldname": f["fieldname"],
						"label": f["label"],
						"parent_field": sub["parent_field"],
						"is_child": True,
					}
				)
	return fields


def get_template(party_type, selected_fields=None, data_rows=None, file_type="xlsx"):
	if not selected_fields:
		selected_fields = _flatten_template_fields(party_type)

	section_row, label_row = [], []
	for field in selected_fields:
		section_row.append(field.get("doctype", ""))
		label_row.append(field.get("label") or field.get("fieldname", ""))

	xlsx_data_rows = []
	if data_rows:
		for row in data_rows:
			xlsx_data_rows.append(
				[cstr(row.get(f"{f['doctype']}::{f['fieldname']}", "")) for f in selected_fields]
			)

	file_type = (file_type or "xlsx").lower()
	if file_type == "csv":
		content = _build_csv(section_row, label_row, xlsx_data_rows)
		filename = f"Party Import Template - {party_type}.csv"
	else:
		content = _build_xlsx(section_row, label_row, xlsx_data_rows, party_type)
		filename = f"Party Import Template - {party_type}.xlsx"

	return content, filename


def _build_csv(section_row, label_row, data_rows):
	buf = io.StringIO()
	writer = csv.writer(buf)
	writer.writerow(section_row)
	writer.writerow(label_row)
	for row in data_rows:
		writer.writerow(row)
	return ("﻿" + buf.getvalue()).encode("utf-8")


def _build_xlsx(section_row, label_row, data_rows, sheet_title):
	try:
		import openpyxl
		from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
		from openpyxl.utils import get_column_letter
	except ImportError:
		frappe.throw(_("openpyxl is required. Run: bench pip install openpyxl"))

	SECTION_DARK = {"Customer": "1F4E79", "Supplier": "1F4E79", "Address": "375623", "Contact": "843C0C"}
	SECTION_LIGHT = {"Customer": "BDD7EE", "Supplier": "BDD7EE", "Address": "C6EFCE", "Contact": "FCE4D6"}
	FALLBACK_DARK, FALLBACK_LIGHT = "404040", "DDDDDD"
	DATA_FILL = PatternFill("solid", fgColor="FFFFFF")
	thin = Side(style="thin", color="D9D9D9")
	border = Border(left=thin, right=thin, bottom=thin)

	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = (sheet_title or "Party Import")[:31]

	for c_idx, (section, label) in enumerate(zip(section_row, label_row, strict=False), start=1):
		dark = SECTION_DARK.get(section, FALLBACK_DARK)
		light = SECTION_LIGHT.get(section, FALLBACK_LIGHT)

		c1 = ws.cell(row=1, column=c_idx, value=section)
		c1.fill = PatternFill("solid", fgColor=dark)
		c1.font = Font(bold=True, color="FFFFFF", size=9)
		c1.alignment = Alignment(horizontal="center", vertical="center")
		c1.border = border

		c2 = ws.cell(row=2, column=c_idx, value=label)
		c2.fill = PatternFill("solid", fgColor=light)
		c2.font = Font(bold=True, size=9)
		c2.alignment = Alignment(wrap_text=True, vertical="center")
		c2.border = border

	for r_idx, row_values in enumerate(data_rows, start=3):
		for c_idx, value in enumerate(row_values, start=1):
			cell = ws.cell(row=r_idx, column=c_idx, value=value)
			cell.fill = DATA_FILL
			cell.font = Font(size=9)
			cell.border = border

	_merge_section_headers(ws, section_row)
	ws.row_dimensions[1].height = 18
	ws.row_dimensions[2].height = 30
	ws.freeze_panes = "A3"

	for c_idx, label in enumerate(label_row, start=1):
		ws.column_dimensions[get_column_letter(c_idx)].width = min(max(len(cstr(label)) + 4, 12), 32)

	buf = io.BytesIO()
	wb.save(buf)
	return buf.getvalue()


def _merge_section_headers(ws, section_row):
	if not section_row:
		return
	start, current = 1, section_row[0]
	for i, val in enumerate(section_row[1:], start=2):
		if val != current:
			if i - 1 > start:
				ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=i - 1)
			start, current = i, val
	end = len(section_row)
	if end >= start:
		ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)


def get_export_data(party_type, selected_fields, limit=None):
	party_doctype = "Supplier" if party_type == "Supplier" else "Customer"
	party_name_field = "customer_name" if party_doctype == "Customer" else "supplier_name"

	if not selected_fields:
		selected_fields = _flatten_template_fields(party_type)

	fields_by_dt = {}
	for f in selected_fields:
		dt, fn = f.get("doctype"), f.get("fieldname")
		if not dt or not fn or fn == "party_code" or f.get("is_child"):
			continue
		fields_by_dt.setdefault(dt, [])
		if fn not in fields_by_dt[dt]:
			fields_by_dt[dt].append(fn)

	party_fields = list({"name", party_name_field, *fields_by_dt.get(party_doctype, [])})
	parties = frappe.get_all(party_doctype, fields=party_fields, limit=limit, order_by="creation asc")
	if not parties:
		return []

	party_names = [p.name for p in parties]

	addr_links = frappe.get_all(
		"Dynamic Link",
		filters={"link_doctype": party_doctype, "link_name": ("in", party_names), "parenttype": "Address"},
		fields=["parent", "link_name"],
	)
	addr_names_by_party = {}
	for lnk in addr_links:
		addr_names_by_party.setdefault(lnk.link_name, []).append(lnk.parent)

	addresses_by_party = {}
	all_addr_names = [lnk.parent for lnk in addr_links]
	if all_addr_names:
		addr_fields = list(
			set(["name"] + (fields_by_dt.get("Address") or ["address_line1", "city", "country"]))
		)
		addr_map = {
			a.name: a
			for a in frappe.get_all("Address", filters={"name": ("in", all_addr_names)}, fields=addr_fields)
		}
		for pname, names in addr_names_by_party.items():
			addresses_by_party[pname] = [addr_map[n] for n in names if n in addr_map]

	cont_links = frappe.get_all(
		"Dynamic Link",
		filters={"link_doctype": party_doctype, "link_name": ("in", party_names), "parenttype": "Contact"},
		fields=["parent", "link_name"],
	)
	cont_names_by_party = {}
	for lnk in cont_links:
		cont_names_by_party.setdefault(lnk.link_name, []).append(lnk.parent)

	contacts_by_party = {}
	all_cont_names = [lnk.parent for lnk in cont_links]
	if all_cont_names:
		cont_fields = list(set(["name"] + (fields_by_dt.get("Contact") or ["first_name", "last_name"])))
		cont_map = {
			c.name: c
			for c in frappe.get_all("Contact", filters={"name": ("in", all_cont_names)}, fields=cont_fields)
		}
		for pname, names in cont_names_by_party.items():
			contacts_by_party[pname] = [cont_map[n] for n in names if n in cont_map]

	from itertools import zip_longest

	rows = []
	for party in parties:
		pname = party.get("name")
		pairs = list(zip_longest(addresses_by_party.get(pname, []), contacts_by_party.get(pname, []))) or [
			(None, None)
		]
		for addr, cont in pairs:
			row = {f"{party_doctype}::party_code": pname}
			for fn in fields_by_dt.get(party_doctype, []):
				row[f"{party_doctype}::{fn}"] = cstr(party.get(fn) or "")
			for fn in fields_by_dt.get("Address", []):
				row[f"Address::{fn}"] = cstr((addr or {}).get(fn) or "")
			for fn in fields_by_dt.get("Contact", []):
				row[f"Contact::{fn}"] = cstr((cont or {}).get(fn) or "")
			rows.append(row)

	return rows


class ImportFile:
	HEADER_ROWS = 2

	def __init__(self, file_url, party_type, validate_only=False):
		self.file_url = file_url
		self.party_type = party_type
		self.validate_only = validate_only
		self.warnings = []
		self.rows = []
		self.col_meta = {}
		self.total_data_rows = 0
		self._parse()

	def _parse(self):
		raw = self._read_raw()
		self._validate_structure(raw)

		section_row = [cstr(v).strip() for v in raw[0]]
		label_row = [cstr(v).strip() for v in raw[1]]
		data_rows = raw[self.HEADER_ROWS :]

		current_section = None
		filled_section_row = []

		for section in section_row:
			if section:
				current_section = section

			filled_section_row.append(current_section)

		column_fieldnames = [
			_fieldname_for_label(dt, lbl) for dt, lbl in zip(filled_section_row, label_row, strict=False)
		]
		column_doctypes = list(filled_section_row)
		for fieldname, doctype in zip(column_fieldnames, column_doctypes, strict=False):
			if fieldname and doctype:
				self.col_meta[f"{doctype}::{fieldname}"] = True

		self.total_data_rows = sum(1 for row in data_rows if any(cstr(v).strip() for v in row))

		if self.validate_only:
			return

		unresolved = [
			f"{dt}::{lbl}"
			for dt, lbl, fn in zip(filled_section_row, label_row, column_fieldnames, strict=False)
			if dt and lbl and not fn
		]
		if unresolved:
			self.warnings.append(
				_("These columns were ignored because no matching field was found: {0}").format(
					", ".join(unresolved)
				)
			)

		party_doctype = "Supplier" if self.party_type == "Supplier" else "Customer"
		if f"{party_doctype}::party_code" not in self.col_meta:
			self.warnings.append(_("Required column 'Party Code' is missing from the file."))

		for i, row in enumerate(data_rows):
			if len(row) < len(column_fieldnames):
				row = [*row, *([""] * (len(column_fieldnames) - len(row)))]

			row_dict = {}
			for fieldname, doctype, val in zip(column_fieldnames, column_doctypes, row, strict=False):
				if not fieldname or not doctype:
					continue
				key = f"{doctype}::{fieldname}"
				row_dict[key] = self._cast_value(fieldname, doctype, cstr(val).strip())

			if not any(row_dict.values()):
				continue

			row_dict["_row_index"] = i + self.HEADER_ROWS + 1
			self.rows.append(row_dict)

	def _validate_structure(self, raw):
		if not raw:
			frappe.throw(_("The uploaded file is empty."), frappe.ValidationError)

		if len(raw) < self.HEADER_ROWS:
			frappe.throw(
				_(
					"Invalid file: expected {0} header rows (Doctype and Label) but found {1}. Download a fresh template and try again."
				).format(self.HEADER_ROWS, len(raw)),
				frappe.ValidationError,
			)

		if len(raw) < self.HEADER_ROWS + 1:
			frappe.throw(
				_("No data rows found. Add at least one data row below the headers."), frappe.ValidationError
			)

	def _read_raw(self):
		file_doc = frappe.get_doc("File", {"file_url": self.file_url})
		content = file_doc.get_content()
		ext = (file_doc.get_extension()[1] or "").lstrip(".").lower()

		if ext == "csv":
			return read_csv_content(content)
		elif ext == "xlsx":
			return read_xlsx_file_from_attached_file(
				fcontent=content if isinstance(content, bytes) else content.encode()
			)
		elif ext == "xls":
			return read_xls_file_from_attached_file(
				content if isinstance(content, bytes) else content.encode()
			)
		else:
			frappe.throw(_("Unsupported file type: .{0}. Upload a CSV, XLS, or XLSX.").format(ext))

	def _cast_value(self, fieldname, doctype, value):
		if not value or not cstr(value).strip():
			return value

		df = frappe.get_meta(doctype).get_field(fieldname) if doctype else None
		if not df:
			return cstr(value)

		ft = df.fieldtype
		try:
			if ft in ("Int", "Check"):
				return cint(value)
			if ft in ("Float", "Currency", "Percent"):
				return flt(value)
			if ft == "Date":
				return getdate(value)
			if ft == "Datetime":
				return get_datetime(value)
			return cstr(value)
		except Exception:
			return cstr(value)

	def get_party_groups(self):
		groups = {}
		order = []
		party_doctype = "Supplier" if self.party_type == "Supplier" else "Customer"
		party_code_key = f"{party_doctype}::party_code"
		name_field = "customer_name" if self.party_type == "Customer" else "supplier_name"
		name_key = f"{party_doctype}::{name_field}"

		for row in self.rows:
			pc = row.get(party_code_key) or row.get(name_key)

			if not pc:
				key = f"_blank_{row.get('_row_index', 0)}"
				g = PartyGroup(key, party_doctype)
				g.add_row(row)
				order.append(key)
				groups[key] = g
				continue

			if pc not in groups:
				groups[pc] = PartyGroup(pc, party_doctype)
				order.append(pc)

			groups[pc].add_row(row)

		return [groups[pc] for pc in order]


def get_or_create_link(doctype, value, auto_create=False, parent_field=None, parent_value=None):
	if not value:
		return None

	if frappe.db.exists(doctype, value):
		return value

	if not auto_create:
		frappe.throw(
			_(
				"{0} {1} does not exist. Enable Create {0} if Missing on the import document, or create it manually first."
			).format(frappe.bold(doctype), frappe.bold(value))
		)

	name_field_map = {
		"Customer Group": "customer_group_name",
		"Supplier Group": "supplier_group_name",
		"Territory": "territory_name",
		"Country": "country_name",
	}

	doc = frappe.new_doc(doctype)
	doc.set(name_field_map.get(doctype, frappe.scrub(doctype) + "_name"), value)

	if parent_field and parent_value:
		doc.set(parent_field, parent_value)
	elif doctype in ("Customer Group", "Supplier Group", "Territory"):
		pf = "parent_" + frappe.scrub(doctype)
		root = frappe.db.get_value(doctype, {"is_group": 1, pf: ("in", ["", None])}, "name")
		if root:
			doc.set(pf, root)

	doc.flags.ignore_permissions = True
	doc.insert()
	return value


class PartyImporter:
	def __init__(self, doc):
		self.doc = doc
		self.party_type = doc.party_type

		self.auto_cg = cint(doc.get("create_customer_group"))
		self.auto_sg = cint(doc.get("create_supplier_group"))
		self.auto_ter = cint(doc.get("create_territory"))

		self.default_cg = doc.get("default_customer_group")
		self.default_sg = doc.get("default_supplier_group")
		self.default_ter = doc.get("default_territory")
		self.default_address_type = doc.get("default_address_type") or "Billing"
		self.default_ct = doc.get("default_customer_type") or "Company"
		self.default_st = doc.get("default_supplier_type") or "Company"

		self.created_parties = {}
		self.contact_cache = {}
		self.log_entries = []
		self.total = self.success = self.failed = 0

	def run(self):
		imf = ImportFile(self.doc.import_file, self.party_type)

		for w in imf.warnings:
			frappe.msgprint(w, indicator="orange", alert=True)

		groups = imf.get_party_groups()
		self.total = len(groups)
		self.doc.db_set("total_records", self.total)

		for idx, group in enumerate(groups):
			frappe.db.savepoint("party_import_group")
			try:
				self._process_group(group)
			except Exception as e:
				frappe.db.rollback(save_point="party_import_group")
				self.failed += 1
				self._log_group(group, "Error", message=cstr(e))
			self._publish_progress(idx + 1)

		self._finalise()

	def _process_group(self, group):
		party_row = group.party_row or {}
		name_field = "customer_name" if self.party_type == "Customer" else "supplier_name"
		party_name = party_row.get(name_field)

		if group.party_code.startswith("_blank_"):
			self.failed += 1
			self._log_group(group, "Error", message=_("party_code is empty — row skipped."))
			return

		party_docname, party_created = self._process_party(group)

		errors = []

		address_created = False
		for addr_row in group.address_rows:
			frappe.db.savepoint("party_import_address")
			try:
				self._process_address(addr_row, party_docname, party_name)
				address_created = True
			except Exception as e:
				frappe.db.rollback(save_point="party_import_address")
				errors.append(_("Address: {0}").format(cstr(e)))

		contact_created = False
		for cont_row in group.contact_rows:
			frappe.db.savepoint("party_import_contact")
			try:
				self._process_contact(cont_row, party_docname)
				contact_created = True
			except Exception as e:
				frappe.db.rollback(save_point="party_import_contact")
				errors.append(_("Contact: {0}").format(cstr(e)))

		if errors:
			self.failed += 1
			self._log_group(
				group,
				"Error",
				docname=party_docname,
				party_created=party_created,
				address_created=address_created,
				contact_created=contact_created,
				message="; ".join(errors),
			)
		else:
			self.success += 1
			self._log_group(
				group,
				"Success",
				docname=party_docname,
				party_created=party_created,
				address_created=address_created,
				contact_created=contact_created,
			)

	def _process_party(self, group):
		party_row = group.party_row or {}
		name_field = "customer_name" if self.party_type == "Customer" else "supplier_name"
		party_name = party_row.get(name_field)

		if group.party_code in self.created_parties:
			return self.created_parties[group.party_code], False

		existing = self._find_existing(group.party_code, party_name)
		if existing:
			self.created_parties[group.party_code] = existing
			return existing, False

		docname = self._create_party(party_row, group.party_code, party_name)
		self.created_parties[group.party_code] = docname
		return docname, True

	def _create_party(self, row, party_code, party_name):
		doc = frappe.new_doc(self.party_type)

		if party_code:
			doc.name = party_code
			doc.flags.name_set = True

		if self.party_type == "Customer":
			ct = row.get("customer_type") or self.default_ct
			cg = (
				row.get("customer_group")
				or self.default_cg
				or frappe.db.get_value("Selling Settings", None, "customer_group")
				or ""
			)
			ter = (
				row.get("territory")
				or self.default_ter
				or frappe.db.get_value("Selling Settings", None, "territory")
				or ""
			)
			doc.customer_name = party_name
			doc.customer_type = ct
			doc.customer_group = (
				get_or_create_link(
					"Customer Group", cg, self.auto_cg, "parent_customer_group", "All Customer Groups"
				)
				or ""
			)
			doc.territory = (
				get_or_create_link("Territory", ter, self.auto_ter, "parent_territory", "All Territories")
				or ""
			)
		else:
			st = row.get("supplier_type") or self.default_st
			sg = (
				row.get("supplier_group")
				or self.default_sg
				or frappe.db.get_value("Buying Settings", None, "supplier_group")
				or ""
			)
			doc.supplier_name = party_name
			doc.supplier_type = st
			doc.supplier_group = (
				get_or_create_link(
					"Supplier Group", sg, self.auto_sg, "parent_supplier_group", "All Supplier Groups"
				)
				or ""
			)

		for fieldname in ("default_currency", "tax_id", "website"):
			if row.get(fieldname):
				doc.set(fieldname, row[fieldname])

		doc.flags.ignore_permissions = True
		doc.insert()
		return doc.name

	def _process_address(self, row, party_docname, party_name):
		addr = frappe.new_doc("Address")
		addr.address_title = party_name or party_docname
		addr.address_type = row.get("address_type") or self.default_address_type
		addr.address_line1 = row.get("address_line1", "")
		addr.address_line2 = row.get("address_line2", "")
		addr.city = row.get("city", "")
		addr.state = row.get("state", "")
		addr.pincode = row.get("pincode", "")
		addr.phone = row.get("phone", "")
		addr.email_id = row.get("email_id", "")
		addr.country = row.get("country") or frappe.db.get_default("country") or "India"
		addr.append("links", {"link_doctype": self.party_type, "link_name": party_docname})
		addr.flags.ignore_permissions = True
		addr.insert()

	def _process_contact(self, row, party_docname):
		first_name = row.get("first_name") or party_docname.split()[0]
		last_name = row.get("last_name", "") or ""

		emails = []
		if row.get("email_id"):
			emails.append(row["email_id"])
		emails.extend(row.get("_extra_emails") or [])

		phones = []
		if row.get("phone"):
			phones.append(row["phone"])
		phones.extend(row.get("_extra_phones") or [])

		existing_name = self._find_existing_contact(party_docname, first_name, last_name)
		if existing_name:
			self._merge_contact_channels(existing_name, emails, phones)
			return

		contact = frappe.new_doc("Contact")
		contact.first_name = first_name
		contact.last_name = last_name
		if row.get("designation"):
			contact.designation = row["designation"]

		seen = set()
		for i, e in enumerate([e for e in emails if e and not (e in seen or seen.add(e))]):
			contact.append("email_ids", {"email_id": e, "is_primary": int(i == 0)})

		seen = set()
		for i, p in enumerate([p for p in phones if p and not (p in seen or seen.add(p))]):
			contact.append("phone_nos", {"phone": p, "is_primary_mobile_no": int(i == 0)})

		contact.append("links", {"link_doctype": self.party_type, "link_name": party_docname})
		contact.flags.ignore_permissions = True
		contact.insert()

		self.contact_cache[(party_docname, first_name, last_name)] = contact.name

	def _find_existing_contact(self, party_docname, first_name, last_name):
		if not first_name:
			return None
		key = (party_docname, first_name, last_name or "")
		if key in self.contact_cache:
			return self.contact_cache[key]

		existing = frappe.db.sql(
			"""
			SELECT c.name FROM `tabContact` c
			INNER JOIN `tabDynamic Link` dl
				ON dl.parent = c.name AND dl.parenttype = 'Contact'
			WHERE dl.link_doctype = %s AND dl.link_name = %s
			  AND c.first_name = %s AND IFNULL(c.last_name, '') = %s
			LIMIT 1
			""",
			(self.party_type, party_docname, first_name, last_name or ""),
			pluck=True,
		)
		if existing:
			self.contact_cache[key] = existing[0]
			return existing[0]
		return None

	def _merge_contact_channels(self, contact_name, new_emails, new_phones):
		contact = frappe.get_doc("Contact", contact_name)
		changed = False

		existing_emails = {e.email_id for e in contact.email_ids if e.email_id}
		has_primary_email = any(e.is_primary for e in contact.email_ids)
		for e in new_emails:
			if not e or e in existing_emails:
				continue
			existing_emails.add(e)
			is_primary = int(not has_primary_email)
			if is_primary:
				has_primary_email = True
			contact.append("email_ids", {"email_id": e, "is_primary": is_primary})
			changed = True

		existing_phones = {p.phone for p in contact.phone_nos if p.phone}
		has_primary_phone = any(p.is_primary_mobile_no for p in contact.phone_nos)
		for p in new_phones:
			if not p or p in existing_phones:
				continue
			existing_phones.add(p)
			is_primary = int(not has_primary_phone)
			if is_primary:
				has_primary_phone = True
			contact.append("phone_nos", {"phone": p, "is_primary_mobile_no": is_primary})
			changed = True

		if changed:
			contact.flags.ignore_permissions = True
			contact.save()

	def _find_existing(self, party_code, party_name):
		if party_code and frappe.db.exists(self.party_type, party_code):
			return party_code
		name_field = "customer_name" if self.party_type == "Customer" else "supplier_name"
		return frappe.db.get_value(self.party_type, {name_field: party_name}, "name")

	def _log_group(
		self,
		group,
		status,
		docname="",
		party_created=False,
		address_created=False,
		contact_created=False,
		message="",
	):
		row = group.party_row or {}
		name_field = "customer_name" if self.party_type == "Customer" else "supplier_name"
		self.log_entries.append(
			{
				"row_index": group.row_indices[0] if group.row_indices else 0,
				"party_name": row.get(name_field) or group.party_code,
				"party_code": group.party_code,
				"status": status,
				"docname": docname,
				"party_created": int(party_created),
				"address_created": int(address_created),
				"contact_created": int(contact_created),
				"message": message,
			}
		)

	def _publish_progress(self, processed):
		frappe.publish_realtime(
			"party_import_progress",
			{
				"party_import": self.doc.name,
				"total": self.total,
				"processed": processed,
				"success": self.success,
				"failed": self.failed,
			},
			doctype="Party Import",
			docname=self.doc.name,
		)

	def _finalise(self):
		if self.failed > 0 and self.success == 0:
			status = "Error"
		elif self.failed > 0:
			status = "Partial Success"
		else:
			status = "Success"

		self.doc.db_set("status", status)
		self.doc.db_set("total_records", self.total)
		self.doc.db_set("success_count", self.success)
		self.doc.db_set("failed_count", self.failed)

		frappe.db.delete("Party Import Log", {"parent": self.doc.name})
		for entry in self.log_entries:
			row = self.doc.append("import_log", entry)
			row.db_insert()

		self.doc.notify_update()


def build_errored_rows_file(doc):
	errored_indices = {
		cint(log_row.row_index): log_row.message or ""
		for log_row in doc.import_log
		if log_row.status in ("Error", "Skipped")
	}

	if not errored_indices:
		frappe.throw(_("No errored or skipped rows found in the import log."))

	imf = ImportFile.__new__(ImportFile)
	imf.file_url = doc.import_file
	imf.party_type = doc.party_type
	raw = imf._read_raw()

	HEADER_ROWS = ImportFile.HEADER_ROWS
	if len(raw) < HEADER_ROWS:
		frappe.throw(_("Original file no longer has valid header rows."))

	header_rows = raw[:HEADER_ROWS]
	data_rows = raw[HEADER_ROWS:]

	errored_data_rows = []
	for i, row in enumerate(data_rows):
		file_row_num = i + HEADER_ROWS + 1
		if file_row_num in errored_indices:
			errored_data_rows.append([*list(row), errored_indices[file_row_num]])

	if not errored_data_rows:
		frappe.throw(
			_(
				"Could not match errored rows to the original file. The file may have changed since the import was run."
			)
		)

	output_header_rows = []
	for i, hrow in enumerate(header_rows):
		hrow = list(hrow)
		hrow.append("" if i == 0 else "Error Message")
		output_header_rows.append(hrow)

	all_rows = output_header_rows + errored_data_rows
	ext = doc.import_file.rsplit(".", 1)[-1].lower()
	label = frappe.scrub(doc.party_type)

	if ext == "csv":
		return _build_errored_csv(all_rows), f"Party Import Errors - {label}.csv"
	return _build_errored_xlsx(all_rows, doc.party_type), f"Party Import Errors - {label}.xlsx"


def _build_errored_csv(all_rows):
	buf = io.StringIO()
	writer = csv.writer(buf)
	for row in all_rows:
		writer.writerow([cstr(v) for v in row])
	return ("﻿" + buf.getvalue()).encode("utf-8")


def _build_errored_xlsx(all_rows, party_type):
	try:
		import openpyxl
		from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
		from openpyxl.utils import get_column_letter
	except ImportError:
		frappe.throw(_("openpyxl is required. Run: bench pip install openpyxl"))

	HEADER_ROWS = ImportFile.HEADER_ROWS
	section_row = [cstr(v) for v in (all_rows[0] if all_rows else [])]
	label_row = [cstr(v) for v in (all_rows[1] if len(all_rows) > 1 else [])]
	data_rows = all_rows[HEADER_ROWS:]

	SECTION_DARK = {"Customer": "1F4E79", "Supplier": "1F4E79", "Address": "375623", "Contact": "843C0C"}
	SECTION_LIGHT = {"Customer": "BDD7EE", "Supplier": "BDD7EE", "Address": "C6EFCE", "Contact": "FCE4D6"}
	FALLBACK_DARK, FALLBACK_LIGHT = "404040", "DDDDDD"
	ERROR_DARK = PatternFill("solid", fgColor="9C0006")
	ERROR_LIGHT = PatternFill("solid", fgColor="FFC7CE")
	DATA_FILL = PatternFill("solid", fgColor="FFFFFF")
	ERROR_DATA = PatternFill("solid", fgColor="FFF2F2")
	thin = Side(style="thin", color="D9D9D9")
	border = Border(left=thin, right=thin, bottom=thin)

	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = f"Errors - {party_type}"[:31]
	n_cols = len(section_row)

	for c_idx, (section, label) in enumerate(zip(section_row, label_row, strict=False), start=1):
		is_err = c_idx == n_cols
		dark = SECTION_DARK.get(section, FALLBACK_DARK)
		light = SECTION_LIGHT.get(section, FALLBACK_LIGHT)

		c1 = ws.cell(row=1, column=c_idx, value="" if is_err else section)
		c1.fill = ERROR_DARK if is_err else PatternFill("solid", fgColor=dark)
		c1.font = Font(bold=True, color="FFFFFF", size=9)
		c1.alignment = Alignment(horizontal="center", vertical="center")
		c1.border = border

		c2 = ws.cell(row=2, column=c_idx, value=label)
		c2.fill = ERROR_LIGHT if is_err else PatternFill("solid", fgColor=light)
		c2.font = Font(bold=True, size=9, color="9C0006" if is_err else "000000")
		c2.alignment = Alignment(wrap_text=True, vertical="center")
		c2.border = border

	_merge_section_headers(ws, section_row)

	for r_idx, row_values in enumerate(data_rows, start=HEADER_ROWS + 1):
		for c_idx, value in enumerate(row_values, start=1):
			is_err = c_idx == n_cols
			cell = ws.cell(row=r_idx, column=c_idx, value=cstr(value))
			cell.fill = ERROR_DATA if is_err else DATA_FILL
			cell.font = Font(size=9, color="9C0006" if is_err else "000000", italic=is_err)
			cell.border = border

	ws.row_dimensions[1].height = 18
	ws.row_dimensions[2].height = 30
	ws.freeze_panes = "A3"

	for c_idx, label in enumerate(label_row, start=1):
		is_err = c_idx == n_cols
		ws.column_dimensions[get_column_letter(c_idx)].width = (
			45 if is_err else min(max(len(cstr(label)) + 4, 12), 32)
		)

	buf = io.BytesIO()
	wb.save(buf)
	return buf.getvalue()
