# Copyright (c) 2019, Frappe Technologies and contributors
# For license information, please see license.txt


import csv
import json
import re

import frappe
import openpyxl
from frappe import _
from frappe.core.doctype.data_import.data_import import DataImport
from frappe.core.doctype.data_import.importer import Importer, ImportFile
from frappe.utils.background_jobs import enqueue
from frappe.utils.xlsxutils import ILLEGAL_CHARACTERS_RE, handle_html
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

INVALID_VALUES = ("", None)


class BankStatementImport(DataImport):
	# begin: auto-generated types
	# This code is auto-generated. Do not modify anything in this block.

	from typing import TYPE_CHECKING

	if TYPE_CHECKING:
		from frappe.types import DF

		bank: DF.Link | None
		bank_account: DF.Link
		company: DF.Link
		custom_delimiters: DF.Check
		delimiter_options: DF.Data | None
		google_sheets_url: DF.Data | None
		import_file: DF.Attach | None
		import_type: DF.Literal["", "Insert New Records", "Update Existing Records"]
		mute_emails: DF.Check
		reference_doctype: DF.Link
		show_failed_logs: DF.Check
		status: DF.Literal["Pending", "Success", "Partial Success", "Error"]
		submit_after_import: DF.Check
		template_options: DF.Code | None
		template_warnings: DF.Code | None
	# end: auto-generated types

	def __init__(self, *args, **kwargs):
		super().__init__(*args, **kwargs)

	def validate(self):
		doc_before_save = self.get_doc_before_save()
		if (
			not (self.import_file or self.google_sheets_url)
			or (doc_before_save and doc_before_save.import_file != self.import_file)
			or (doc_before_save and doc_before_save.google_sheets_url != self.google_sheets_url)
		):
			template_options_dict = {}
			column_to_field_map = {}
			bank = frappe.get_doc("Bank", self.bank)
			for i in bank.bank_transaction_mapping:
				column_to_field_map[i.file_field] = i.bank_transaction_field
			template_options_dict["column_to_field_map"] = column_to_field_map
			self.template_options = json.dumps(template_options_dict)

			self.template_warnings = ""

		self.validate_import_file()
		self.validate_google_sheets_url()

	def start_import(self):
		preview = frappe.get_doc("Bank Statement Import", self.name).get_preview_from_template(
			self.import_file, self.google_sheets_url
		)

		if "Bank Account" not in json.dumps(preview["columns"]):
			frappe.throw(_("Please add the Bank Account column"))

		from frappe.utils.background_jobs import is_job_enqueued
		from frappe.utils.scheduler import is_scheduler_inactive

		if is_scheduler_inactive() and not frappe.flags.in_test:
			frappe.throw(_("Scheduler is inactive. Cannot import data."), title=_("Scheduler Inactive"))

		job_id = f"bank_statement_import::{self.name}"
		if not is_job_enqueued(job_id):
			enqueue(
				start_import,
				queue="default",
				timeout=6000,
				event="data_import",
				job_id=job_id,
				data_import=self.name,
				bank_account=self.bank_account,
				import_file_path=self.import_file,
				google_sheets_url=self.google_sheets_url,
				bank=self.bank,
				template_options=self.template_options,
				now=frappe.conf.developer_mode or frappe.flags.in_test,
			)
			return True

		return False


@frappe.whitelist()
def get_preview_from_template(data_import, import_file=None, google_sheets_url=None):
	return frappe.get_doc("Bank Statement Import", data_import).get_preview_from_template(
		import_file, google_sheets_url
	)


@frappe.whitelist()
def form_start_import(data_import):
	return frappe.get_doc("Bank Statement Import", data_import).start_import()


@frappe.whitelist()
def download_errored_template(data_import_name):
	data_import = frappe.get_doc("Bank Statement Import", data_import_name)
	data_import.export_errored_rows()


@frappe.whitelist()
def download_import_log(data_import_name):
	return frappe.get_doc("Bank Statement Import", data_import_name).download_import_log()


def parse_data_from_template(raw_data):
	data = []

	for _i, row in enumerate(raw_data):
		if all(v in INVALID_VALUES for v in row):
			# empty row
			continue

		data.append(row)

	return data


def start_import(data_import, bank_account, import_file_path, google_sheets_url, bank, template_options):
	"""This method runs in background job"""

	update_mapping_db(bank, template_options)

	data_import = frappe.get_doc("Bank Statement Import", data_import)
	file = import_file_path if import_file_path else google_sheets_url

	import_file = ImportFile("Bank Transaction", file=file, import_type="Insert New Records")

	data = parse_data_from_template(import_file.raw_data)
	# Importer expects 'Data Import' class, which has 'payload_count' attribute
	if not data_import.get("payload_count"):
		data_import.payload_count = len(data) - 1

	if import_file_path:
		add_bank_account(data, bank_account)
		write_files(import_file, data)

	try:
		i = Importer(data_import.reference_doctype, data_import=data_import)
		i.import_data()
	except Exception:
		frappe.db.rollback()
		data_import.db_set("status", "Error")
		data_import.log_error("Bank Statement Import failed")
	finally:
		frappe.flags.in_import = False

	frappe.publish_realtime("data_import_refresh", {"data_import": data_import.name})


def update_mapping_db(bank, template_options):
	bank = frappe.get_doc("Bank", bank)
	for d in bank.bank_transaction_mapping:
		d.delete()

	for d in json.loads(template_options)["column_to_field_map"].items():
		bank.append("bank_transaction_mapping", {"bank_transaction_field": d[1], "file_field": d[0]})

	bank.save()


def add_bank_account(data, bank_account):
	bank_account_loc = None
	if "Bank Account" not in data[0]:
		data[0].append("Bank Account")
	else:
		for loc, header in enumerate(data[0]):
			if header == "Bank Account":
				bank_account_loc = loc

	for row in data[1:]:
		if bank_account_loc:
			row[bank_account_loc] = bank_account
		else:
			row.append(bank_account)


def write_files(import_file, data):
	full_file_path = import_file.file_doc.get_full_path()
	parts = import_file.file_doc.get_extension()
	extension = parts[1]
	extension = extension.lstrip(".")

	if extension == "csv":
		with open(full_file_path, "w", newline="") as file:
			writer = csv.writer(file)
			writer.writerows(data)
	elif extension == "xlsx" or "xls":
		write_xlsx(data, "trans", file_path=full_file_path)


def write_xlsx(data, sheet_name, wb=None, column_widths=None, file_path=None):
	# from xlsx utils with changes
	column_widths = column_widths or []
	if wb is None:
		wb = openpyxl.Workbook(write_only=True)

	ws = wb.create_sheet(sheet_name, 0)

	for i, column_width in enumerate(column_widths):
		if column_width:
			ws.column_dimensions[get_column_letter(i + 1)].width = column_width

	row1 = ws.row_dimensions[1]
	row1.font = Font(name="Calibri", bold=True)

	for row in data:
		clean_row = []
		for item in row:
			if isinstance(item, str) and (sheet_name not in ["Data Import Template", "Data Export"]):
				value = handle_html(item)
			else:
				value = item

			if isinstance(item, str) and next(ILLEGAL_CHARACTERS_RE.finditer(value), None):
				# Remove illegal characters from the string
				value = re.sub(ILLEGAL_CHARACTERS_RE, "", value)

			clean_row.append(value)

		ws.append(clean_row)

	wb.save(file_path)
	return True


@frappe.whitelist()
def get_import_status(docname):
	import_status = {}

	data_import = frappe.get_doc("Bank Statement Import", docname)
	import_status["status"] = data_import.status

	logs = frappe.get_all(
		"Data Import Log",
		fields=["count(*) as count", "success"],
		filters={"data_import": docname},
		group_by="success",
	)

	total_payload_count = 0

	for log in logs:
		total_payload_count += log.get("count", 0)
		if log.get("success"):
			import_status["success"] = log.get("count")
		else:
			import_status["failed"] = log.get("count")

	import_status["total_records"] = total_payload_count

	return import_status


@frappe.whitelist()
def get_import_logs(docname: str):
	frappe.has_permission("Bank Statement Import")

	return frappe.get_all(
		"Data Import Log",
		fields=["success", "docname", "messages", "exception", "row_indexes"],
		filters={"data_import": docname},
		limit_page_length=5000,
		order_by="log_index",
	)


@frappe.whitelist()
def upload_bank_statement(**args):
	args = frappe._dict(args)
	bsi = frappe.new_doc("Bank Statement Import")

	if args.company:
		bsi.update(
			{
				"company": args.company,
			}
		)

	if args.bank_account:
		bsi.update({"bank_account": args.bank_account})

	return bsi
