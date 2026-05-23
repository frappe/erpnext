# Copyright (c) 2026, Frappe Technologies Pvt. Ltd. and contributors
# For license information, please see license.txt

"""Reads CSV and XLSX uploads into a list-of-dict row stream.

The reader is forgiving about CSV encoding (Excel-on-Windows exports are
common), and it surfaces non-fatal observations (e.g. multi-sheet workbooks)
through a caller-provided warnings list rather than raising or printing.
"""

import csv
import os

import frappe
from frappe import _
from frappe.utils import cstr
from frappe.utils.file_manager import get_file_path

CSV_ENCODINGS = ("utf-8-sig", "utf-8", "cp1252", "latin-1")
CSV_EXTENSIONS = (".csv",)
EXCEL_EXTENSIONS = (".xlsx", ".xls")


class FileReader:
	"""Reads a Frappe File-URL into rows keyed by header name."""

	def __init__(self, file_url: str, warnings: list[str] | None = None):
		self.file_url = file_url
		self.warnings = warnings

	def read(self) -> list[dict]:
		"""Return rows as list of dicts; raises frappe.ValidationError on unknown type."""
		file_path = get_file_path(self.file_url)
		extension = os.path.splitext(file_path)[1].lower()

		if extension in CSV_EXTENSIONS:
			return self._read_csv(file_path)
		if extension in EXCEL_EXTENSIONS:
			return self._read_xlsx(file_path)
		frappe.throw(_("Unsupported file type: {0}").format(extension))

	def _read_csv(self, file_path: str) -> list[dict]:
		last_error = None
		for encoding in CSV_ENCODINGS:
			try:
				with open(file_path, encoding=encoding, newline="") as f:
					return [dict(r) for r in csv.DictReader(f)]
			except UnicodeDecodeError as exc:
				last_error = exc
		self._raise_decode_error(last_error)

	def _read_xlsx(self, file_path: str) -> list[dict]:
		workbook = self._load_workbook(file_path)
		self._warn_if_multi_sheet(workbook)
		worksheet = workbook.active
		row_iterator = worksheet.iter_rows(values_only=True)
		headers = [cstr(h) for h in next(row_iterator, [])]
		return list(self._iter_xlsx_rows(row_iterator, headers))

	def _load_workbook(self, file_path: str):
		try:
			from openpyxl import load_workbook
		except ImportError:
			frappe.throw(_("openpyxl is required to read Excel files"))
		return load_workbook(filename=file_path, read_only=True, data_only=True)

	def _warn_if_multi_sheet(self, workbook) -> None:
		if self.warnings is None or len(workbook.sheetnames) <= 1:
			return
		other = [s for s in workbook.sheetnames if s != workbook.active.title]
		preview = ", ".join(other[:3]) + ("…" if len(other) > 3 else "")
		self.warnings.append(
			_(
				"Excel file has {0} sheets. Only the first sheet (<b>{1}</b>) was imported. "
				"Other sheets ignored: {2}."
			).format(len(workbook.sheetnames), workbook.active.title, preview)
		)

	def _iter_xlsx_rows(self, row_iterator, headers: list[str]):
		for row in row_iterator:
			if not any(c is not None and cstr(c).strip() for c in row):
				continue
			yield {
				headers[i]: cstr(c) if c is not None else "" for i, c in enumerate(row) if i < len(headers)
			}

	def _raise_decode_error(self, last_error) -> None:
		message = _(
			"Could not decode the CSV file. Save it as UTF-8 from Excel "
			"(File → Save As → CSV UTF-8) and try again."
		)
		if last_error:
			message += f"<br><br><code>{last_error}</code>"
		frappe.throw(message)
